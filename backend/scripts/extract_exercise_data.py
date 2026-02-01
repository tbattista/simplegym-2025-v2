"""
Extract Exercise Data from Excel
Matches the 190 kept exercises against the full Excel database
and extracts all columns including video URLs.
"""

import json
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook, Workbook


def normalize_name(name: str) -> str:
    """Normalize exercise name for comparison"""
    if not name:
        return ""
    return name.lower().strip()


def get_hyperlink_or_value(cell):
    """Extract hyperlink URL if present, otherwise return cell value"""
    if cell.hyperlink:
        return cell.hyperlink.target
    return cell.value


def load_exercises_to_keep(file_path: Path) -> list:
    """Load the list of exercises to keep"""
    with open(file_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    return data.get('exercises', [])


def find_matching_row(exercise_name: str, excel_data: list, name_col_idx: int) -> tuple:
    """
    Find a matching row in Excel data for the given exercise name.
    Returns (row_data, match_type) if found, (None, None) otherwise.
    """
    normalized_search = normalize_name(exercise_name)
    if not normalized_search:
        return None, None

    # First pass: exact match
    for row in excel_data:
        excel_name = row.get(name_col_idx, '')
        if excel_name and normalize_name(str(excel_name)) == normalized_search:
            return row, "exact"

    # Second pass: partial match (search term in excel name)
    for row in excel_data:
        excel_name = row.get(name_col_idx, '')
        if excel_name:
            normalized_excel = normalize_name(str(excel_name))
            if normalized_search in normalized_excel:
                return row, "partial"

    # Third pass: partial match (excel name in search term)
    for row in excel_data:
        excel_name = row.get(name_col_idx, '')
        if excel_name:
            normalized_excel = normalize_name(str(excel_name))
            if len(normalized_excel) > 5 and normalized_excel in normalized_search:
                return row, "contains"

    return None, None


def main():
    print("\n" + "="*80)
    print("EXTRACT EXERCISE DATA FROM EXCEL")
    print("="*80)

    # Paths
    script_dir = Path(__file__).parent
    project_root = script_dir.parent.parent

    exercises_to_keep_path = script_dir / "analysis_results" / "exercises_to_keep.json"
    excel_path = project_root / "full exercise db rip.xlsx"
    output_dir = script_dir / "analysis_results"

    # Check files exist
    if not exercises_to_keep_path.exists():
        print(f"Error: {exercises_to_keep_path} not found")
        return

    if not excel_path.exists():
        print(f"Error: {excel_path} not found")
        return

    # Load exercises to keep
    print(f"\nLoading exercises to keep from: {exercises_to_keep_path}")
    exercises_to_keep = load_exercises_to_keep(exercises_to_keep_path)
    print(f"Found {len(exercises_to_keep)} exercises to match")

    # Load Excel file
    print(f"\nLoading Excel file: {excel_path}")
    wb = load_workbook(excel_path)
    ws = wb.active

    # Find header row (contains "Exercise" in column B)
    HEADER_ROW = 16  # Based on inspection
    DATA_START_ROW = 17
    NAME_COL = 2  # Column B (1-indexed)

    # Get headers from row 16
    headers = []
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=HEADER_ROW, column=col)
        val = get_hyperlink_or_value(cell)
        headers.append(val if val else f"Column_{col}")

    print(f"\nHeaders (row {HEADER_ROW}):")
    for i, h in enumerate(headers[:15]):
        print(f"  Col {i+1}: {h}")
    if len(headers) > 15:
        print(f"  ... and {len(headers) - 15} more")

    # Load all Excel data into memory (with hyperlinks extracted)
    print(f"\nLoading Excel data from row {DATA_START_ROW}...")
    excel_data = []
    for row_num in range(DATA_START_ROW, ws.max_row + 1):
        row_data = {}
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_num, column=col)
            row_data[col] = get_hyperlink_or_value(cell)
        # Only add rows that have an exercise name
        if row_data.get(NAME_COL):
            excel_data.append(row_data)

    print(f"Loaded {len(excel_data)} exercise rows from Excel")

    # Show sample Excel exercises
    print("\nSample exercises from Excel:")
    for i, row in enumerate(excel_data[:10]):
        print(f"  {i+1}. {row.get(NAME_COL)}")

    # Match exercises
    print("\n" + "-"*40)
    print("Matching exercises...")
    print("-"*40)

    matched = []
    unmatched = []

    for exercise in exercises_to_keep:
        name = exercise.get('name', '')
        match, match_type = find_matching_row(name, excel_data, NAME_COL)

        if match:
            matched.append({
                'firebase_id': exercise.get('id'),
                'firebase_name': name,
                'match_type': match_type,
                'excel_name': match.get(NAME_COL),
                'excel_data': match
            })
            print(f"  MATCHED ({match_type}): {name} -> {match.get(NAME_COL)}")
        else:
            unmatched.append(exercise)

    print(f"\n  Total matched: {len(matched)}")
    print(f"  Total unmatched: {len(unmatched)}")

    # Create output Excel file
    print("\n" + "-"*40)
    print("Creating output files...")
    print("-"*40)

    output_wb = Workbook()
    output_ws = output_wb.active
    output_ws.title = "Matched Exercises"

    # Write headers
    output_headers = ['Firebase_ID', 'Firebase_Name', 'Match_Type', 'Excel_Name'] + headers[1:]  # Skip empty col A
    for col, header in enumerate(output_headers, 1):
        output_ws.cell(row=1, column=col, value=header)

    # Write matched data
    for row_num, match in enumerate(matched, 2):
        output_ws.cell(row=row_num, column=1, value=match['firebase_id'])
        output_ws.cell(row=row_num, column=2, value=match['firebase_name'])
        output_ws.cell(row=row_num, column=3, value=match['match_type'])
        output_ws.cell(row=row_num, column=4, value=match['excel_name'])

        # Write Excel data (columns B onwards)
        for col_num in range(2, ws.max_column + 1):
            value = match['excel_data'].get(col_num)
            output_ws.cell(row=row_num, column=col_num + 3, value=value)

    # Save output Excel
    output_excel_path = output_dir / "matched_exercises_with_videos.xlsx"
    output_wb.save(output_excel_path)
    print(f"Saved matched exercises to: {output_excel_path}")

    # Save unmatched to JSON
    unmatched_path = output_dir / "unmatched_exercises.json"
    with open(unmatched_path, 'w', encoding='utf-8') as f:
        json.dump({
            'count': len(unmatched),
            'generated': datetime.now().isoformat(),
            'exercises': unmatched
        }, f, indent=2)
    print(f"Saved unmatched exercises to: {unmatched_path}")

    # Summary
    print("\n" + "="*80)
    print("SUMMARY")
    print("="*80)
    print(f"Total exercises to keep: {len(exercises_to_keep)}")
    print(f"Matched in Excel: {len(matched)} ({100*len(matched)/len(exercises_to_keep):.1f}%)")
    print(f"Not found in Excel: {len(unmatched)} ({100*len(unmatched)/len(exercises_to_keep):.1f}%)")

    if matched:
        print("\nFirst 10 matched exercises:")
        for m in matched[:10]:
            print(f"  - {m['firebase_name']} -> {m['excel_name']} ({m['match_type']})")

    if unmatched:
        print("\nFirst 20 unmatched exercises:")
        for ex in unmatched[:20]:
            print(f"  - {ex['name']}")
        if len(unmatched) > 20:
            print(f"  ... and {len(unmatched) - 20} more")


if __name__ == "__main__":
    main()
