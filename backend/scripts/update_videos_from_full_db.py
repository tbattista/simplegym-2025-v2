"""
Update Firebase Exercises with Video URLs from Full Exercise DB
Reads 'full exercise db rip.xlsx' and pushes shortVideoUrl + detailedVideoUrl to Firestore
by matching exercise names (case-insensitive exact match).

Usage:
    python backend/scripts/update_videos_from_full_db.py --dry-run
    python backend/scripts/update_videos_from_full_db.py
"""

import sys
import json
import argparse
from pathlib import Path
from datetime import datetime, timezone

from openpyxl import load_workbook

project_root = Path(__file__).parent.parent.parent
sys.path.insert(0, str(project_root))

from dotenv import load_dotenv
load_dotenv(project_root / '.env')

from backend.config.firebase_config import get_firebase_app
from firebase_admin import firestore


EXCEL_PATH = project_root / "full exercise db rip.xlsx"
DB_DUMP_PATH = project_root / "backend" / "scripts" / "analysis_results" / "db_complete_dump.json"
HEADER_ROW = 16
NAME_COL = 2
SHORT_COL = 3
DETAILED_COL = 4


def load_excel_videos():
    """Load exercises with video hyperlinks from the full exercise DB Excel."""
    print(f"Loading: {EXCEL_PATH}")
    wb = load_workbook(EXCEL_PATH)
    ws = wb['Exercises']

    exercises = []
    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        name = ws.cell(row=row, column=NAME_COL).value
        if not name:
            continue

        short_cell = ws.cell(row=row, column=SHORT_COL)
        detailed_cell = ws.cell(row=row, column=DETAILED_COL)

        short_url = short_cell.hyperlink.target if short_cell.hyperlink else None
        detailed_url = detailed_cell.hyperlink.target if detailed_cell.hyperlink else None

        if short_url or detailed_url:
            videos = {}
            if short_url:
                videos['shortVideoUrl'] = short_url.strip()
            if detailed_url:
                videos['detailedVideoUrl'] = detailed_url.strip()
            exercises.append({'name': name.strip(), 'videos': videos})

    return exercises


def build_name_index_from_dump():
    """Build name->doc_id index from local db dump (avoids Firestore pagination issues)."""
    print(f"Loading name index from: {DB_DUMP_PATH}")
    with open(DB_DUMP_PATH, 'r') as f:
        db_data = json.load(f)

    exercises_list = db_data.get('global_exercises', [])
    index = {}
    for ex in exercises_list:
        name = (ex.get('name') or '').strip().lower()
        eid = ex.get('id', '')
        if name and eid:
            index[name] = eid
    print(f"  Indexed {len(index)} exercises from dump")
    return index


def match_and_update(excel_exercises, name_index, db, dry_run=False):
    """Match Excel exercises to Firestore by name and push video URLs."""
    mode = "DRY RUN" if dry_run else "LIVE"
    print(f"\n{'='*60}")
    print(f"UPDATE FIRESTORE ({mode})")
    print(f"{'='*60}\n")

    matched = []
    unmatched = []

    for ex in excel_exercises:
        key = ex['name'].lower()
        if key in name_index:
            matched.append({
                'doc_id': name_index[key],
                'name': ex['name'],
                'videos': ex['videos']
            })
        else:
            unmatched.append(ex['name'])

    print(f"Matched: {len(matched)}")
    print(f"Unmatched: {len(unmatched)}")

    if dry_run:
        short_count = sum(1 for m in matched if 'shortVideoUrl' in m['videos'])
        detailed_count = sum(1 for m in matched if 'detailedVideoUrl' in m['videos'])
        print(f"\n  Short video URLs: {short_count}")
        print(f"  Detailed video URLs: {detailed_count}")
        print(f"\nWould update {len(matched)} documents")

        if unmatched:
            print(f"\nFirst 20 unmatched:")
            for name in unmatched[:20]:
                print(f"  - {name}")
        return len(matched)

    # Live update using batched writes (max 500 per batch)
    BATCH_SIZE = 400
    updated = 0
    failed = 0

    for i in range(0, len(matched), BATCH_SIZE):
        batch = db.batch()
        chunk = matched[i:i + BATCH_SIZE]

        for m in chunk:
            doc_ref = db.collection('global_exercises').document(m['doc_id'])
            batch.update(doc_ref, m['videos'])

        try:
            batch.commit()
            updated += len(chunk)
            print(f"  Batch {i // BATCH_SIZE + 1}: updated {len(chunk)} exercises ({updated}/{len(matched)} total)")
        except Exception as e:
            failed += len(chunk)
            print(f"  Batch {i // BATCH_SIZE + 1}: FAILED — {e}")

    # Summary
    short_count = sum(1 for m in matched if 'shortVideoUrl' in m['videos'])
    detailed_count = sum(1 for m in matched if 'detailedVideoUrl' in m['videos'])

    print(f"\n{'='*60}")
    print(f"SUMMARY")
    print(f"{'='*60}")
    print(f"Total matched:           {len(matched)}")
    print(f"  Short video URLs:      {short_count}")
    print(f"  Detailed video URLs:   {detailed_count}")
    print(f"Updated:                 {updated}")
    if failed:
        print(f"Failed:                  {failed}")
    print(f"Unmatched (no Firestore): {len(unmatched)}")

    return updated


def main():
    parser = argparse.ArgumentParser(description='Push video URLs from full exercise DB to Firestore')
    parser.add_argument('--dry-run', action='store_true', help='Preview without writing')
    args = parser.parse_args()

    if not EXCEL_PATH.exists():
        print(f"ERROR: {EXCEL_PATH} not found")
        return

    excel_exercises = load_excel_videos()
    print(f"Found {len(excel_exercises)} exercises with video URLs in Excel")

    app = get_firebase_app()
    if not app:
        print("ERROR: Failed to initialize Firebase")
        return

    db = firestore.client(app=app)
    name_index = build_name_index_from_dump()

    updated = match_and_update(excel_exercises, name_index, db, dry_run=args.dry_run)

    # Bump metadata version if live update
    if not args.dry_run and updated > 0:
        try:
            ref = db.collection('exercises_metadata').document('global')
            doc = ref.get()
            if doc.exists:
                data = doc.to_dict()
                old_ver = data.get('version', '1.0.0')
                parts = old_ver.split('.')
                parts[-1] = str(int(parts[-1]) + 1)
                new_ver = '.'.join(parts)
            else:
                new_ver = '1.0.1'
            ref.set({'version': new_ver, 'lastUpdated': datetime.now(timezone.utc).isoformat()}, merge=True)
            print(f"\nMetadata version bumped to {new_ver}")
        except Exception as e:
            print(f"\nWARNING: Failed to bump metadata version: {e}")


if __name__ == "__main__":
    main()
