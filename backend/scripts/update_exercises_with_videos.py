"""
Update Firebase Exercises with Video URLs
Reads matched_exercises_with_videos.xlsx and pushes shortVideoUrl + detailedVideoUrl to Firestore.

Usage:
    python backend/scripts/update_exercises_with_videos.py --dry-run
    python backend/scripts/update_exercises_with_videos.py
"""

import sys
import argparse
from pathlib import Path
from datetime import datetime

from openpyxl import load_workbook

# Add project root to path for imports
project_root = Path(__file__).parent.parent.parent
sys.path.insert(0, str(project_root))

from dotenv import load_dotenv
load_dotenv(project_root / '.env')

from backend.config.firebase_config import get_firebase_app
from firebase_admin import firestore


EXCEL_PATH = Path(__file__).parent / "analysis_results" / "matched_exercises_with_videos.xlsx"

# Only video URL fields — other metadata is already in Firestore
FIELD_MAPPING = {
    'Short YouTube Demonstration': 'shortVideoUrl',
    'In-Depth YouTube Explanation': 'detailedVideoUrl',
}


def load_exercises():
    """Load exercises from Excel, extracting hyperlink targets for video URLs."""
    print(f"Loading: {EXCEL_PATH}")
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active

    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    col_idx = {h: i + 1 for i, h in enumerate(headers)}

    exercises = []
    for row in range(2, ws.max_row + 1):
        fid = ws.cell(row=row, column=col_idx['Firebase_ID']).value
        if not fid:
            continue

        name = ws.cell(row=row, column=col_idx['Firebase_Name']).value

        # Extract video URLs (prefer hyperlink target over cell value)
        videos = {}
        for excel_col, firebase_field in FIELD_MAPPING.items():
            ci = col_idx.get(excel_col)
            if not ci:
                continue
            cell = ws.cell(row=row, column=ci)
            url = cell.hyperlink.target if cell.hyperlink else cell.value
            if url and isinstance(url, str) and url.strip():
                videos[firebase_field] = url.strip()

        if videos:
            exercises.append({'id': fid, 'name': name, 'videos': videos})

    return exercises


def update_firestore(exercises, dry_run=False):
    """Push video URLs to Firestore global_exercises collection."""
    mode = "DRY RUN" if dry_run else "LIVE"
    print(f"\n{'='*60}")
    print(f"UPDATE FIRESTORE ({mode}) — {len(exercises)} exercises with videos")
    print(f"{'='*60}\n")

    app = get_firebase_app()
    if not app:
        print("ERROR: Failed to initialize Firebase")
        return False

    db = firestore.client(app=app)
    updated = 0
    not_found = []

    for i, ex in enumerate(exercises, 1):
        doc_ref = db.collection('global_exercises').document(ex['id'])

        if dry_run:
            # Verify document exists
            doc = doc_ref.get()
            exists = doc.exists
            status = "OK" if exists else "NOT FOUND"
            if not exists:
                not_found.append(ex)
            fields = ', '.join(f"{k}" for k in ex['videos'])
            print(f"  [{i}/{len(exercises)}] {status} | {ex['name']} | {fields}")
            updated += 1 if exists else 0
        else:
            try:
                doc_ref.update(ex['videos'])
                print(f"  [{i}/{len(exercises)}] Updated: {ex['name']}")
                updated += 1
            except Exception as e:
                print(f"  [{i}/{len(exercises)}] FAILED: {ex['name']} — {e}")

    # Summary
    short_count = sum(1 for ex in exercises if 'shortVideoUrl' in ex['videos'])
    detailed_count = sum(1 for ex in exercises if 'detailedVideoUrl' in ex['videos'])

    print(f"\n{'='*60}")
    print(f"SUMMARY")
    print(f"{'='*60}")
    print(f"Total with videos:       {len(exercises)}")
    print(f"  Short video URLs:      {short_count}")
    print(f"  Detailed video URLs:   {detailed_count}")
    print(f"{'Updated' if not dry_run else 'Would update'}: {updated}")
    if not_found:
        print(f"Not found in Firestore:  {len(not_found)}")
        for ex in not_found:
            print(f"    - {ex['name']} ({ex['id']})")

    return len(not_found) == 0


def main():
    parser = argparse.ArgumentParser(description='Push video URLs to Firestore exercises')
    parser.add_argument('--dry-run', action='store_true', help='Preview without writing')
    args = parser.parse_args()

    if not EXCEL_PATH.exists():
        print(f"ERROR: {EXCEL_PATH} not found")
        return

    exercises = load_exercises()
    print(f"Found {len(exercises)} exercises with video URLs")

    if not exercises:
        print("Nothing to update")
        return

    update_firestore(exercises, dry_run=args.dry_run)


if __name__ == "__main__":
    main()
