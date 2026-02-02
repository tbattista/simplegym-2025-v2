"""
Update Firebase Exercises with Video URLs
Reads the matched exercises Excel file and updates Firebase with video URLs and metadata.
"""

import sys
import json
from pathlib import Path
from datetime import datetime
from dotenv import load_dotenv
from openpyxl import load_workbook

# Load environment variables
env_path = Path(__file__).parent.parent.parent / '.env'
load_dotenv(env_path)

# Add parent directory to path for imports
sys.path.append(str(Path(__file__).parent.parent.parent))

from backend.config.firebase_config import get_firebase_app
from firebase_admin import firestore


def load_matched_exercises(file_path: Path) -> list:
    """Load matched exercises from Excel file"""
    print(f"Loading matched exercises from: {file_path}")

    wb = load_workbook(file_path)
    ws = wb.active

    # Get headers from first row
    headers = []
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        headers.append(cell.value)

    print(f"Found {len(headers)} columns")

    # Map column indices to header names
    col_map = {h: i for i, h in enumerate(headers)}

    # Load data rows
    exercises = []
    for row_num in range(2, ws.max_row + 1):
        row_data = {}
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_num, column=col)
            header = headers[col - 1]
            row_data[header] = cell.value

        # Only add if has Firebase ID
        if row_data.get('Firebase_ID'):
            exercises.append(row_data)

    print(f"Loaded {len(exercises)} exercises")
    return exercises, headers


def update_exercises(exercises: list, dry_run: bool = False):
    """Update exercises in Firebase"""
    print("\n" + "="*80)
    print(f"UPDATE EXERCISES IN FIREBASE" + (" (DRY RUN)" if dry_run else ""))
    print("="*80)

    if dry_run:
        print("\n*** DRY RUN MODE - No changes will be made ***\n")

    # Initialize Firebase
    app = get_firebase_app()
    if not app:
        print("Failed to initialize Firebase")
        return False

    db = firestore.client(app=app)
    print("Connected to Firestore\n")

    # Field mapping from Excel headers to Firebase field names
    field_mapping = {
        'Short YouTube Demonstration': 'shortVideoUrl',
        'In-Depth YouTube Explanation': 'detailedVideoUrl',
        'Difficulty Level': 'difficultyLevel',
        'Target Muscle Group ': 'targetMuscleGroup',  # Note: has trailing space in Excel
        'Target Muscle Group': 'targetMuscleGroup',
        'Prime Mover Muscle': 'primeMoverMuscle',
        'Secondary Muscle': 'secondaryMuscle',
        'Tertiary Muscle': 'tertiaryMuscle',
        'Primary Equipment ': 'primaryEquipment',  # Note: has trailing space in Excel
        'Primary Equipment': 'primaryEquipment',
        '# Primary Items': 'primaryEquipmentCount',
        'Secondary Equipment': 'secondaryEquipment',
        '# Secondary Items': 'secondaryEquipmentCount',
        'Posture': 'posture',
        'Single or Double Arm': 'armType',
        'Continuous or Alternating': 'armPattern',
        'Grip': 'grip',
        'Load Position (end of movement)': 'loadPosition',
        'Foot Elevation': 'footElevation',
        'Single Exercise, Combination, or Combo Rep': 'combinationExercise',
        'Movement Pattern 1': 'movementPattern1',
        'Movement Pattern 2': 'movementPattern2',
        'Movement Pattern 3': 'movementPattern3',
        'Plane of Motion 1': 'planeOfMotion1',
        'Plane of Motion 2': 'planeOfMotion2',
        'Plane of Motion 3': 'planeOfMotion3',
        'Body Region': 'bodyRegion',
        'Force': 'forceType',
        'Mechanics': 'mechanics',
        'Laterality': 'laterality',
        'Primary Classification': 'classification',
    }

    updated_count = 0
    failed = []
    update_log = []

    for i, exercise in enumerate(exercises):
        firebase_id = exercise.get('Firebase_ID')
        firebase_name = exercise.get('Firebase_Name')

        if not firebase_id:
            continue

        # Build update data
        update_data = {}

        for excel_field, firebase_field in field_mapping.items():
            value = exercise.get(excel_field)
            if value is not None and value != '':
                # Clean up the value
                if isinstance(value, str):
                    value = value.strip()
                update_data[firebase_field] = value

        # Add updated timestamp
        update_data['updatedAt'] = datetime.now().isoformat()

        # Count how many video URLs we're adding
        has_short_video = bool(update_data.get('shortVideoUrl'))
        has_detailed_video = bool(update_data.get('detailedVideoUrl'))

        if dry_run:
            print(f"  [{i+1}/{len(exercises)}] Would update: {firebase_name}")
            print(f"      Short Video: {'Yes' if has_short_video else 'No'}")
            print(f"      Detailed Video: {'Yes' if has_detailed_video else 'No'}")
            print(f"      Fields to update: {len(update_data)}")
            update_log.append({
                'id': firebase_id,
                'name': firebase_name,
                'action': 'would_update',
                'fields': list(update_data.keys()),
                'has_short_video': has_short_video,
                'has_detailed_video': has_detailed_video
            })
            updated_count += 1
        else:
            try:
                doc_ref = db.collection('global_exercises').document(firebase_id)
                doc_ref.update(update_data)
                print(f"  [{i+1}/{len(exercises)}] Updated: {firebase_name}")
                update_log.append({
                    'id': firebase_id,
                    'name': firebase_name,
                    'action': 'updated',
                    'fields': list(update_data.keys()),
                    'has_short_video': has_short_video,
                    'has_detailed_video': has_detailed_video
                })
                updated_count += 1
            except Exception as e:
                print(f"  [{i+1}/{len(exercises)}] FAILED: {firebase_name} - {e}")
                failed.append({
                    'id': firebase_id,
                    'name': firebase_name,
                    'error': str(e)
                })

    # Save update log
    output_dir = Path(__file__).parent / "analysis_results"
    log_file = output_dir / ("update_log_dry_run.json" if dry_run else "update_log.json")

    with open(log_file, 'w', encoding='utf-8') as f:
        json.dump({
            'dry_run': dry_run,
            'timestamp': datetime.now().isoformat(),
            'total_processed': len(exercises),
            'updated_count': updated_count,
            'failed_count': len(failed),
            'updates': update_log,
            'failures': failed
        }, f, indent=2)

    print(f"\nUpdate log saved: {log_file}")

    # Summary
    print("\n" + "="*80)
    print("SUMMARY")
    print("="*80)
    print(f"Total exercises processed: {len(exercises)}")
    print(f"Successfully {'would update' if dry_run else 'updated'}: {updated_count}")
    if failed:
        print(f"Failed: {len(failed)}")

    # Count videos
    with_short_video = sum(1 for u in update_log if u.get('has_short_video'))
    with_detailed_video = sum(1 for u in update_log if u.get('has_detailed_video'))
    print(f"\nExercises with short video URL: {with_short_video}")
    print(f"Exercises with detailed video URL: {with_detailed_video}")

    if dry_run:
        print("\n*** This was a DRY RUN - No changes were made ***")
        print("Run without --dry-run to execute updates")

    return len(failed) == 0


def main():
    import argparse

    parser = argparse.ArgumentParser(description='Update exercises with video URLs')
    parser.add_argument('--dry-run', action='store_true',
                        help='Preview updates without actually updating')

    args = parser.parse_args()

    # Paths
    script_dir = Path(__file__).parent
    excel_path = script_dir / "analysis_results" / "matched_exercises_with_videos.xlsx"

    if not excel_path.exists():
        print(f"Error: {excel_path} not found")
        print("Run extract_exercise_data.py first")
        return

    # Load exercises
    exercises, headers = load_matched_exercises(excel_path)

    if not exercises:
        print("No exercises to update")
        return

    # Update exercises
    success = update_exercises(exercises, dry_run=args.dry_run)

    if success:
        print("\nUpdate process completed successfully")
    else:
        print("\nUpdate process completed with errors")


if __name__ == "__main__":
    main()
